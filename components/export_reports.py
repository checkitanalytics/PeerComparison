import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from fpdf import FPDF

def render_export_options(comparison_data):
    """
    Render export options for downloading comparison reports
    
    Args:
        comparison_data (dict): Dictionary containing financial data for each company
    """
    if not comparison_data:
        return
    
    st.sidebar.markdown("---")
    st.sidebar.subheader("📥 Export Reports")
    
    export_format = st.sidebar.radio(
        "Select Export Format",
        options=["Excel (.xlsx)", "PDF Report"],
        help="Choose the format for your comparison report"
    )
    
    if st.sidebar.button("📥 Download Report", type="primary"):
        with st.spinner("Generating report..."):
            if export_format == "Excel (.xlsx)":
                excel_file = generate_excel_report(comparison_data)
                if excel_file:
                    st.sidebar.download_button(
                        label="⬇️ Download Excel Report",
                        data=excel_file,
                        file_name=f"company_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            else:
                pdf_file = generate_pdf_report(comparison_data)
                if pdf_file:
                    st.sidebar.download_button(
                        label="⬇️ Download PDF Report",
                        data=pdf_file,
                        file_name=f"company_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf"
                    )

def generate_excel_report(comparison_data):
    """
    Generate Excel report with multiple sheets for different metrics
    
    Args:
        comparison_data (dict): Dictionary containing financial data for each company
    
    Returns:
        bytes: Excel file as bytes
    """
    try:
        output = BytesIO()
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Overview sheet
        ws_overview = wb.create_sheet("Overview")
        create_overview_sheet(ws_overview, comparison_data)
        
        # Create Valuation sheet
        ws_valuation = wb.create_sheet("Valuation Metrics")
        create_valuation_sheet(ws_valuation, comparison_data)
        
        # Create Profitability sheet
        ws_profitability = wb.create_sheet("Profitability")
        create_profitability_sheet(ws_profitability, comparison_data)
        
        # Create Financial Performance sheet
        ws_performance = wb.create_sheet("Financial Performance")
        create_performance_sheet(ws_performance, comparison_data)
        
        # Create Financial Health sheet
        ws_health = wb.create_sheet("Financial Health")
        create_health_sheet(ws_health, comparison_data)
        
        # Save workbook
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Excel report: {str(e)}")
        return None

def create_overview_sheet(ws, comparison_data):
    """Create overview sheet with basic company information"""
    # Header
    ws['A1'] = 'Company Comparison Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    
    # Create header row
    headers = ['Company', 'Sector', 'Industry', 'Market Cap', 'Current Price', 'P/E Ratio', '52W High', '52W Low']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    row = 5
    for company, data in comparison_data.items():
        ws.cell(row=row, column=1, value=company)
        ws.cell(row=row, column=2, value=data.get('Sector', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('Industry', 'N/A'))
        ws.cell(row=row, column=4, value=data.get('Market Cap', 'N/A'))
        ws.cell(row=row, column=5, value=data.get('Current Price', 'N/A'))
        ws.cell(row=row, column=6, value=data.get('P/E Ratio', 'N/A'))
        ws.cell(row=row, column=7, value=data.get('52 Week High', 'N/A'))
        ws.cell(row=row, column=8, value=data.get('52 Week Low', 'N/A'))
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 18

def create_valuation_sheet(ws, comparison_data):
    """Create valuation metrics sheet"""
    headers = ['Company', 'Market Cap', 'P/E Ratio', 'Forward P/E', 'P/B Ratio', 'EV/EBITDA', 
               'Dividend Yield', 'Dividend Rate']
    
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for company, data in comparison_data.items():
        ws.cell(row=row, column=1, value=company)
        ws.cell(row=row, column=2, value=data.get('Market Cap', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('P/E Ratio', 'N/A'))
        ws.cell(row=row, column=4, value=data.get('Forward P/E', 'N/A'))
        ws.cell(row=row, column=5, value=data.get('P/B Ratio', 'N/A'))
        ws.cell(row=row, column=6, value=data.get('EV/EBITDA', 'N/A'))
        ws.cell(row=row, column=7, value=data.get('Dividend Yield', 'N/A'))
        ws.cell(row=row, column=8, value=data.get('Dividend Rate', 'N/A'))
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 16

def create_profitability_sheet(ws, comparison_data):
    """Create profitability metrics sheet"""
    headers = ['Company', 'Gross Margin', 'Operating Margin', 'Profit Margin', 'ROE', 'ROA',
               'Revenue Growth', 'Earnings Growth']
    
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for company, data in comparison_data.items():
        ws.cell(row=row, column=1, value=company)
        ws.cell(row=row, column=2, value=data.get('Gross Margin', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('Operating Margin', 'N/A'))
        ws.cell(row=row, column=4, value=data.get('Profit Margin', 'N/A'))
        ws.cell(row=row, column=5, value=data.get('ROE', 'N/A'))
        ws.cell(row=row, column=6, value=data.get('ROA', 'N/A'))
        ws.cell(row=row, column=7, value=data.get('Revenue Growth', 'N/A'))
        ws.cell(row=row, column=8, value=data.get('Earnings Growth', 'N/A'))
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        ws.column_dimensions[chr(64 + col)].width = 18

def create_performance_sheet(ws, comparison_data):
    """Create financial performance sheet"""
    headers = ['Company', 'Revenue (TTM)', 'Gross Profit (TTM)', 'Net Income (TTM)', 
               'Operating Cash Flow (TTM)', 'Free Cash Flow (TTM)']
    
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for company, data in comparison_data.items():
        ws.cell(row=row, column=1, value=company)
        ws.cell(row=row, column=2, value=data.get('Revenue (TTM)', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('Gross Profit (TTM)', 'N/A'))
        ws.cell(row=row, column=4, value=data.get('Net Income (TTM)', 'N/A'))
        ws.cell(row=row, column=5, value=data.get('Operating Cash Flow (TTM)', 'N/A'))
        ws.cell(row=row, column=6, value=data.get('Free Cash Flow (TTM)', 'N/A'))
        row += 1
    
    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[chr(64 + col)].width = 22

def create_health_sheet(ws, comparison_data):
    """Create financial health sheet"""
    headers = ['Company', 'Current Ratio', 'Quick Ratio', 'Debt to Equity', 
               'Total Debt', 'Total Cash', 'Payout Ratio']
    
    # Header row
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    row = 2
    for company, data in comparison_data.items():
        ws.cell(row=row, column=1, value=company)
        ws.cell(row=row, column=2, value=data.get('Current Ratio', 'N/A'))
        ws.cell(row=row, column=3, value=data.get('Quick Ratio', 'N/A'))
        ws.cell(row=row, column=4, value=data.get('Debt to Equity', 'N/A'))
        ws.cell(row=row, column=5, value=data.get('Total Debt', 'N/A'))
        ws.cell(row=row, column=6, value=data.get('Total Cash', 'N/A'))
        ws.cell(row=row, column=7, value=data.get('Payout Ratio', 'N/A'))
        row += 1
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 18

def generate_pdf_report(comparison_data):
    """
    Generate PDF report with company comparison
    
    Args:
        comparison_data (dict): Dictionary containing financial data for each company
    
    Returns:
        bytes: PDF file as bytes
    """
    try:
        pdf = FPDF()
        pdf.add_page()
        
        # Title
        pdf.set_font('Arial', 'B', 20)
        pdf.cell(0, 10, 'Company Comparison Report', ln=True, align='C')
        
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 10, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', ln=True, align='C')
        pdf.ln(10)
        
        # Overview section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Company Overview', ln=True)
        pdf.ln(5)
        
        for company, data in comparison_data.items():
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, company, ln=True)
            
            pdf.set_font('Arial', '', 10)
            pdf.cell(0, 6, f"Sector: {data.get('Sector', 'N/A')}", ln=True)
            pdf.cell(0, 6, f"Industry: {data.get('Industry', 'N/A')}", ln=True)
            
            # Format market cap
            market_cap = data.get('Market Cap')
            if pd.notna(market_cap):
                mc_str = f"${market_cap/1e9:.2f}B" if market_cap >= 1e9 else f"${market_cap/1e6:.2f}M"
            else:
                mc_str = "N/A"
            pdf.cell(0, 6, f"Market Cap: {mc_str}", ln=True)
            
            # Format current price
            price = data.get('Current Price')
            price_str = f"${price:.2f}" if pd.notna(price) else "N/A"
            pdf.cell(0, 6, f"Current Price: {price_str}", ln=True)
            
            pdf.ln(3)
        
        # Valuation Metrics
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Valuation Metrics', ln=True)
        pdf.ln(5)
        
        for company, data in comparison_data.items():
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, company, ln=True)
            
            pdf.set_font('Arial', '', 10)
            pe_ratio = data.get('P/E Ratio')
            pdf.cell(0, 6, f"P/E Ratio: {pe_ratio:.2f}" if pd.notna(pe_ratio) else "P/E Ratio: N/A", ln=True)
            
            pb_ratio = data.get('P/B Ratio')
            pdf.cell(0, 6, f"P/B Ratio: {pb_ratio:.2f}" if pd.notna(pb_ratio) else "P/B Ratio: N/A", ln=True)
            
            ev_ebitda = data.get('EV/EBITDA')
            pdf.cell(0, 6, f"EV/EBITDA: {ev_ebitda:.2f}" if pd.notna(ev_ebitda) else "EV/EBITDA: N/A", ln=True)
            
            pdf.ln(3)
        
        # Profitability Metrics
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Profitability Metrics', ln=True)
        pdf.ln(5)
        
        for company, data in comparison_data.items():
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, company, ln=True)
            
            pdf.set_font('Arial', '', 10)
            
            # Format margins as percentages
            gross_margin = data.get('Gross Margin')
            if pd.notna(gross_margin):
                gm_str = f"{gross_margin*100:.2f}%" if gross_margin < 1 else f"{gross_margin:.2f}%"
            else:
                gm_str = "N/A"
            pdf.cell(0, 6, f"Gross Margin: {gm_str}", ln=True)
            
            profit_margin = data.get('Profit Margin')
            if pd.notna(profit_margin):
                pm_str = f"{profit_margin*100:.2f}%" if profit_margin < 1 else f"{profit_margin:.2f}%"
            else:
                pm_str = "N/A"
            pdf.cell(0, 6, f"Profit Margin: {pm_str}", ln=True)
            
            roe = data.get('ROE')
            if pd.notna(roe):
                roe_str = f"{roe*100:.2f}%" if roe < 1 else f"{roe:.2f}%"
            else:
                roe_str = "N/A"
            pdf.cell(0, 6, f"ROE: {roe_str}", ln=True)
            
            pdf.ln(3)
        
        # Financial Performance
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Financial Performance (TTM)', ln=True)
        pdf.ln(5)
        
        for company, data in comparison_data.items():
            pdf.set_font('Arial', 'B', 12)
            pdf.cell(0, 8, company, ln=True)
            
            pdf.set_font('Arial', '', 10)
            
            revenue = data.get('Revenue (TTM)')
            if pd.notna(revenue):
                rev_str = f"${revenue/1e9:.2f}B" if revenue >= 1e9 else f"${revenue/1e6:.2f}M"
            else:
                rev_str = "N/A"
            pdf.cell(0, 6, f"Revenue: {rev_str}", ln=True)
            
            net_income = data.get('Net Income (TTM)')
            if pd.notna(net_income):
                ni_str = f"${net_income/1e9:.2f}B" if abs(net_income) >= 1e9 else f"${net_income/1e6:.2f}M"
            else:
                ni_str = "N/A"
            pdf.cell(0, 6, f"Net Income: {ni_str}", ln=True)
            
            fcf = data.get('Free Cash Flow (TTM)')
            if pd.notna(fcf):
                fcf_str = f"${fcf/1e9:.2f}B" if abs(fcf) >= 1e9 else f"${fcf/1e6:.2f}M"
            else:
                fcf_str = "N/A"
            pdf.cell(0, 6, f"Free Cash Flow: {fcf_str}", ln=True)
            
            pdf.ln(3)
        
        # Footer
        pdf.set_y(-15)
        pdf.set_font('Arial', 'I', 8)
        pdf.cell(0, 10, 'Data provided by Yahoo Finance. This report is for informational purposes only.', align='C')
        
        # Output PDF as bytes
        return pdf.output(dest='S').encode('latin-1')
        
    except Exception as e:
        st.error(f"Error generating PDF report: {str(e)}")
        return None
